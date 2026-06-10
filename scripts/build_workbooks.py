"""Build all working/ and solutions/ .xlsx files from the source CSVs.

Run after `generate_dataset.py`. Re-running is safe — files are overwritten.

UAE FMCG distributor context. One synthetic dataset (sales of Food/HPC brands to
Dubai supermarkets, with Sales and Returns) travels through 5 workflow stages.
File names stay module-N.xlsx, mapped 1:1 to the stages:
- module-1  Stage 1 · Receive & clean   (Table, TRIM/CLEAN, Find&Replace, VALUE/DATEVALUE,
                                          text functions, dedupe, formatting)
- module-2  Stage 2 · Structure & enrich (VLOOKUP SalesRep->Manager/Quota from Reps,
                                          Brand->Brand Manager from Brands; failure modes)
- module-3  Stage 3 · Explore & summarize (AutoFilter, SUBTOTAL; SUMIFS/COUNTIFS/AVERAGEIFS;
                                          net sales, gross, return rate, AOV, quota attainment)
- module-4  Stage 4 · Pivot & rank        (PivotTables/slicers; rep & brand-manager leaderboards)
- module-5  Stage 5 · Present             (KPI one-pager, one chart, conditional formatting)
- capstone  fresh file: run the whole workflow, answer ~11 questions.

No commission anywhere. PivotTables / Subtotal outlines can't be authored by openpyxl,
so solution files mirror them with SUMIFS tables.
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
FILES = ROOT / "docs" / "files"
SOURCE_CSV = FILES / "source" / "sales_data.csv"
REPS_CSV = FILES / "source" / "reps.csv"
BRANDS_CSV = FILES / "source" / "brands.csv"
WORKING_DIR = FILES / "working"
SOLUTIONS_DIR = FILES / "solutions"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(italic=True, color="7F6000")
TASK_FILL = PatternFill("solid", fgColor="E2EFDA")

AED = '"AED" #,##0'          # summary money
AED2 = '#,##0.00'           # line-item money

# Example entities referenced consistently across exercises & lessons.
EG_REP = "Mohammed Saleh"
EG_BRANDMGR = "Imran Sheikh"
EG_AREA = "Deira"
EG_CATEGORY = "Food"
EG_BRAND = "Crunchio"
EG_CUSTOMER = "Carrefour"


# ----------------------------- data loading ---------------------------------

HEADERS = ["OrderNumber", "Date", "InvoiceType", "CustomerCode", "Customer",
           "BranchCode", "Branch", "Area", "SalesRep", "Brand", "Category",
           "SalesQuantity", "SalesValue"]


def load_rows() -> list[dict]:
    with SOURCE_CSV.open(encoding="utf-8") as f:
        rows = []
        for r in csv.DictReader(f):
            r["CustomerCode"] = int(r["CustomerCode"])
            r["BranchCode"] = int(r["BranchCode"])
            r["SalesQuantity"] = int(r["SalesQuantity"])
            r["SalesValue"] = float(r["SalesValue"])
            rows.append(r)
    return rows


def load_reps() -> list[dict]:
    with REPS_CSV.open(encoding="utf-8") as f:
        reps = []
        for r in csv.DictReader(f):
            r["AnnualQuota"] = int(r["AnnualQuota"])
            reps.append(r)
    return reps


def load_brands() -> list[dict]:
    with BRANDS_CSV.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _as_date(iso: str) -> date:
    y, m, d = iso.split("-")
    return date(int(y), int(m), int(d))


def _cleaned_rows(rows: list[dict]) -> list[dict]:
    """Trim Customer, normalise Area casing, drop exact-duplicate order lines."""
    cleaned, seen = [], set()
    for r in rows:
        c = dict(r)
        c["Customer"] = c["Customer"].strip()
        c["Area"] = c["Area"].title()
        c["Branch"] = f'{c["Customer"]} - {c["Area"]}'
        key = (c["Date"], c["InvoiceType"], c["CustomerCode"], c["BranchCode"],
               c["SalesRep"], c["Brand"], c["SalesQuantity"], c["SalesValue"])
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(c)
    return cleaned


# ----------------------------- sheet writers --------------------------------

COL_WIDTHS = {"A": 12, "B": 12, "C": 11, "D": 13, "E": 24, "F": 12, "G": 28,
              "H": 14, "I": 18, "J": 14, "K": 11, "L": 14, "M": 13, "N": 16, "O": 16}


def _style_header(ws, ncols: int) -> None:
    for col_idx in range(1, ncols + 1):
        ws.cell(row=1, column=col_idx).font = HEADER_FONT
        ws.cell(row=1, column=col_idx).fill = HEADER_FILL
    for col_letter, w in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w


def _row_values(r: dict, *, typed_dates: bool) -> list:
    return [
        r["OrderNumber"],
        _as_date(r["Date"]) if typed_dates else r["Date"],
        r["InvoiceType"], r["CustomerCode"], r["Customer"], r["BranchCode"],
        r["Branch"], r["Area"], r["SalesRep"], r["Brand"], r["Category"],
        r["SalesQuantity"], r["SalesValue"],
    ]


def _format_data_rows(ws, last_row: int) -> None:
    for i in range(2, last_row + 1):
        ws.cell(row=i, column=2).number_format = "yyyy-mm-dd"   # Date
        ws.cell(row=i, column=4).number_format = "0"            # CustomerCode
        ws.cell(row=i, column=6).number_format = "0"            # BranchCode
        ws.cell(row=i, column=12).number_format = "#,##0"       # SalesQuantity
        ws.cell(row=i, column=13).number_format = AED2          # SalesValue


def write_sales_table(ws, rows: list[dict], *, table_name: str = "Sales") -> None:
    """Clean data as an Excel Table (13 base columns)."""
    ws.append(HEADERS)
    for r in rows:
        ws.append(_row_values(r, typed_dates=True))
    table = Table(displayName=table_name, ref=f"A1:M{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    _style_header(ws, len(HEADERS))
    _format_data_rows(ws, ws.max_row)
    ws.freeze_panes = "A2"


def write_sales_range(ws, rows: list[dict]) -> None:
    """Plain RAW range for the Stage 1 cleaning exercise.

    Deliberately left un-frozen, with SalesQuantity/SalesValue unformatted:
    freezing the header row (exercise 9) and formatting SalesValue (exercise 10)
    are the student's job. Only Date and the code columns get a display format so
    they don't render as raw serial numbers."""
    ws.append(HEADERS)
    for r in rows:
        ws.append(_row_values(r, typed_dates=True))
    _style_header(ws, len(HEADERS))
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=2).number_format = "yyyy-mm-dd"   # Date
        ws.cell(row=i, column=4).number_format = "0"            # CustomerCode
        ws.cell(row=i, column=6).number_format = "0"            # BranchCode


def write_sales_enriched(ws, rows: list[dict], *, table_name: str = "Sales") -> None:
    """Sales Table + two VLOOKUP'd columns (SalesManager, BrandManager).
    Needs 'Reps' and 'Brands' sheets in the same workbook."""
    headers = list(HEADERS) + ["SalesManager", "BrandManager"]
    ws.append(headers)
    for r in rows:
        ws.append(_row_values(r, typed_dates=True))
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=14, value=f"=VLOOKUP(I{i},Reps[#All],2,FALSE)")
        ws.cell(row=i, column=15, value=f"=VLOOKUP(J{i},Brands[#All],3,FALSE)")
    table = Table(displayName=table_name, ref=f"A1:O{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    _style_header(ws, len(headers))
    _format_data_rows(ws, ws.max_row)
    ws.freeze_panes = "A2"


def write_reps_table(ws, reps: list[dict]) -> None:
    ws.append(["SalesRep", "Manager", "AnnualQuota"])
    for r in reps:
        ws.append([r["SalesRep"], r["Manager"], r["AnnualQuota"]])
    t = Table(displayName="Reps", ref=f"A1:C{ws.max_row}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(t)
    for c in range(1, 4):
        ws.cell(row=1, column=c).font = HEADER_FONT
        ws.cell(row=1, column=c).fill = HEADER_FILL
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=3).number_format = AED
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 15
    ws.freeze_panes = "A2"


def write_brands_table(ws, brands: list[dict]) -> None:
    ws.append(["Brand", "Category", "BrandManager"])
    for b in brands:
        ws.append([b["Brand"], b["Category"], b["BrandManager"]])
    t = Table(displayName="Brands", ref=f"A1:C{ws.max_row}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium5", showRowStripes=True)
    ws.add_table(t)
    for c in range(1, 4):
        ws.cell(row=1, column=c).font = HEADER_FONT
        ws.cell(row=1, column=c).fill = HEADER_FILL
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 16
    ws.freeze_panes = "A2"


def add_refs(wb, reps, brands):
    write_reps_table(wb.create_sheet("Reps"), reps)
    write_brands_table(wb.create_sheet("Brands"), brands)


# ----------------------------- small helpers --------------------------------

def add_note(ws, text: str, *, cell: str = "A1", span: str = "H") -> None:
    ws[cell] = text
    ws[cell].fill = NOTE_FILL
    ws[cell].font = NOTE_FONT
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"{cell}:{span}1")
    ws.row_dimensions[1].height = 30


def section(ws, row: int, text: str) -> None:
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, color="1F4E78", size=12)


def task_header(ws, row: int, cols: list[str], widths: list[int]) -> None:
    for i, (label, w) in enumerate(zip(cols, widths), start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(i)].width = w


def hint(ws, row: int, text: str) -> None:
    ws.cell(row=row, column=2, value=text).font = Font(italic=True, color="808080")


def task_list(ws, items: list[str], *, start: int = 3) -> int:
    for i, t in enumerate(items, start=start):
        section(ws, i, t)
    return start + len(items)


# ===================== Stage 1 · Receive & clean ============================

def build_stage1(rows: list[dict]) -> None:
    # ---- working ----
    wb = Workbook()
    ws = wb.active
    ws.title = "RawData"
    write_sales_range(ws, rows)
    add_note(ws, "RAW export from the system — messy on purpose. Step 1: turn it into a clean "
                 "Excel Table called 'Sales'. Watch for spaces in Customer names, MiXeD-case "
                 "Areas, and duplicate order lines.", cell="O1", span="O")

    clean = wb.create_sheet("Cleanup Practice")
    add_note(clean, "Cleanup Practice — fix each messy value in column B using the function in "
                    "column A; type your formula in column C.")
    task_header(clean, 3, ["Function", "Messy value", "Your fix"], [16, 30, 30])
    practice = [
        ("TRIM", "  Carrefour  ", "text", '=TRIM(B4)'),
        ("CLEAN", '=CHAR(10)&"Lulu Hypermarket"', "formula", '=CLEAN(B5)'),
        ("TRIM+CLEAN", '=CHAR(9)&"  Spinneys  "', "formula", '=TRIM(CLEAN(B6))'),
        ("VALUE", "1250.50", "text", '=VALUE(B7)'),
        ("VALUE", "89", "text", '=VALUE(B8)'),
        ("DATEVALUE", "2024-03-15", "text", '=DATEVALUE(B9)'),
        ("DATEVALUE", "2025-11-02", "text", '=DATEVALUE(B10)'),
        ("SUBSTITUTE", "Deira;Dubai", "text", '=SUBSTITUTE(B11,";"," ")'),
        ("LEFT", "SO-100562", "text", '=LEFT(B12,2)'),
        ("MID", "SO-100562", "text", '=MID(B13,4,6)'),
        ("RIGHT", "SO-100562", "text", '=RIGHT(B14,6)'),
        ('& (combine)', "100562", "text", '="SO-"&B15'),
    ]
    _write_practice(clean, practice, fill=False)
    tip = 4 + len(practice) + 1
    clean.cell(row=tip, column=1, value="Tip:").font = Font(bold=True, color="7F6000")
    hint(clean, tip, "Text-stored numbers/dates sit on the LEFT of the cell; real ones on the right. "
                     "VALUE/DATEVALUE convert them. LEFT/MID/RIGHT split an order number; & joins text.")

    ex = wb.create_sheet("Exercises")
    add_note(ex, "Stage 1 exercises — clean the raw export. See the lesson for full steps.")
    task_list(ex, [
        "1. Select the RawData range and press Ctrl+T to make an Excel Table; name it 'Sales'.",
        "2. TRIM the spaces from Customer names (helper column, then paste values back).",
        "3. CLEAN any hidden line breaks/tabs (see Cleanup Practice).",
        "4. Find & Replace (Ctrl+H): fix CAPS areas — replace 'DEIRA' with 'Deira', etc.",
        "5. Convert text numbers/dates on Cleanup Practice with VALUE and DATEVALUE.",
        "6. Use LEFT/MID/RIGHT and SUBSTITUTE to split an OrderNumber; combine text with &.",
        "7. Paste Special > Values to lock a helper column, then delete the original.",
        "8. Data > Remove Duplicates on the order columns — how many were removed?",
        "9. Multi-sort: Customer (A-Z), then Date (newest first); freeze the header row.",
        "10. Format: SalesValue as number with 2 decimals; try the AED #,##0 custom format.",
    ])
    wb.save(WORKING_DIR / "module-1.xlsx")

    # ---- solution ----
    wb = Workbook()
    rm = wb.active
    rm.title = "ReadMe"
    add_note(rm, "Stage 1 solution. 'Sales' is a clean Excel Table: trimmed Customers, normalised "
                 "Area casing, duplicates removed, sorted. 'Cleanup Practice' shows every fix.")
    ws = wb.create_sheet("Sales")
    cleaned = sorted(_cleaned_rows(rows), key=lambda r: (r["Customer"], r["Date"]))
    write_sales_table(ws, cleaned)
    clean = wb.create_sheet("Cleanup Practice")
    add_note(clean, "Cleanup Practice — solved.")
    task_header(clean, 3, ["Function", "Messy value", "Fixed"], [16, 30, 30])
    _write_practice(clean, practice, fill=True)
    wb.save(SOLUTIONS_DIR / "module-1.xlsx")


def _write_practice(ws, practice, *, fill: bool) -> None:
    for i, (fn, messy, kind, sol) in enumerate(practice, start=4):
        ws.cell(row=i, column=1, value=fn).font = Font(bold=True)
        c = ws.cell(row=i, column=2)
        c.value = messy
        if kind == "text":
            c.data_type = "s"
        if fill:
            out = ws.cell(row=i, column=3, value=sol)
            if fn == "DATEVALUE":
                out.number_format = "yyyy-mm-dd"
            elif fn == "VALUE":
                out.number_format = "#,##0.00"


# ===================== Stage 2 · Structure & enrich =========================

def build_stage2(rows: list[dict], reps: list[dict], brands: list[dict]) -> None:
    cleaned = _cleaned_rows(rows)

    # ---- working ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    write_sales_table(ws, cleaned)
    add_refs(wb, reps, brands)
    ex = wb.create_sheet("Exercises", 0)
    add_note(ex, "Stage 2 — enrich 'Sales' with VLOOKUP from the 'Reps' and 'Brands' tables.")
    task_list(ex, [
        "1. Add a 'SalesManager' column:  =VLOOKUP([@SalesRep], Reps, 2, FALSE).",
        "2. Add a 'Quota' column:         =VLOOKUP([@SalesRep], Reps, 3, FALSE).",
        "3. Add a 'BrandManager' column:  =VLOOKUP([@Brand], Brands, 3, FALSE).",
        "4. Why FALSE? Try TRUE on a rep and watch the answer go wrong.",
        "5. VLOOKUP can't look LEFT: try to fetch SalesRep from Manager — what happens?",
        "6. Wrap a missing lookup in IFERROR(..., \"unknown\") to kill the #N/A.",
        "7. (M365) Redo task 1 with XLOOKUP: =XLOOKUP([@SalesRep], Reps[SalesRep], Reps[Manager], \"unknown\").",
    ])
    wb.save(WORKING_DIR / "module-2.xlsx")

    # ---- solution ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    # enriched with SalesManager + BrandManager + a Quota column
    headers = list(HEADERS) + ["SalesManager", "BrandManager", "Quota"]
    ws.append(headers)
    for r in cleaned:
        ws.append(_row_values(r, typed_dates=True))
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=14, value=f"=VLOOKUP(I{i},Reps[#All],2,FALSE)")
        ws.cell(row=i, column=15, value=f"=VLOOKUP(J{i},Brands[#All],3,FALSE)")
        ws.cell(row=i, column=16, value=f"=VLOOKUP(I{i},Reps[#All],3,FALSE)")
        ws.cell(row=i, column=16).number_format = AED
    t = Table(displayName="Sales", ref=f"A1:{get_column_letter(len(headers))}{ws.max_row}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    _style_header(ws, len(headers))
    ws.column_dimensions["P"].width = 14
    _format_data_rows(ws, ws.max_row)
    ws.freeze_panes = "A2"
    add_refs(wb, reps, brands)

    fm = wb.create_sheet("Failure Modes", 0)
    add_note(fm, "The six classic VLOOKUP failures — and the fix for each.")
    task_header(fm, 3, ["#", "Failure mode", "Fix"], [4, 46, 60])
    failures = [
        ("Approximate match returns wrong row", "Always pass FALSE (exact) as the 4th argument."),
        ("Can't look to the LEFT of the key column",
         "Re-order columns so the key is leftmost, or use INDEX/MATCH / XLOOKUP."),
        ("Column index breaks when a column is inserted",
         "Reference a Table (Reps/Brands) so the index tracks, or use XLOOKUP."),
        ("Number-stored-as-text vs real number mismatch",
         "Make both sides the same type (VALUE, or Text-to-Columns)."),
        ("Trailing/leading spaces in the key", "TRIM both sides first (Stage 1!)."),
        ("#N/A when the value genuinely isn't there", "Wrap in IFERROR(VLOOKUP(...), \"unknown\")."),
    ]
    for i, (mode, fix) in enumerate(failures, start=4):
        fm.cell(row=i, column=1, value=i - 3)
        fm.cell(row=i, column=2, value=mode)
        fm.cell(row=i, column=3, value=fix)
    section(fm, 12, "Modern note (Excel 365 / 2021+)")
    fm["B13"] = "XLOOKUP fixes most of these by default — exact match, any direction, survives inserts:"
    # Shown as an example, not evaluated: [@Brand] only resolves inside the Sales
    # table, so store it as literal text (data_type 's') to avoid a #REF!/#NAME?.
    fm["B14"].value = '=XLOOKUP([@Brand], Brands[Brand], Brands[BrandManager], "unknown")'
    fm["B14"].data_type = "s"
    fm["B14"].font = Font(name="Consolas")
    wb.save(SOLUTIONS_DIR / "module-2.xlsx")


# ===================== Stage 3 · Explore & summarize ========================

STAGE3_PROMPTS: list[tuple[str, str, str]] = [
    ("1. Net sales (sum of SalesValue — returns net out automatically)",
     "=SUM(Sales[SalesValue])", AED),
    ("2. Gross sales (Sales invoices only)",
     '=SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Sales")', AED),
    ("3. Total returns value (Return invoices — a negative number)",
     '=SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Return")', AED),
    ("4. KPI — Return rate (returns / gross, as a positive %)",
     '=-SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Return")'
     '/SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Sales")', "0.0%"),
    ("5. Number of invoice lines", "=COUNTA(Sales[OrderNumber])", "#,##0"),
    ("6. KPI — Average sale value (Sales lines only)",
     '=AVERAGEIFS(Sales[SalesValue],Sales[InvoiceType],"Sales")', AED),
    (f"7. Net sales in the {EG_AREA} area",
     f'=SUMIFS(Sales[SalesValue],Sales[Area],"{EG_AREA}")', AED),
    (f"8. Net sales for the {EG_CATEGORY} category",
     f'=SUMIFS(Sales[SalesValue],Sales[Category],"{EG_CATEGORY}")', AED),
    (f"9. Net sales booked by rep {EG_REP}",
     f'=SUMIFS(Sales[SalesValue],Sales[SalesRep],"{EG_REP}")', AED),
    (f"10. Net sales for brand manager {EG_BRANDMGR} (uses the BrandManager column)",
     f'=SUMIFS(Sales[SalesValue],Sales[BrandManager],"{EG_BRANDMGR}")', AED),
    (f"11. {EG_REP}'s quota attainment (net sales / quota from Reps)",
     f'=SUMIFS(Sales[SalesValue],Sales[SalesRep],"{EG_REP}")/VLOOKUP("{EG_REP}",Reps[#All],3,FALSE)', "0.0%"),
]


def build_stage3(rows: list[dict], reps: list[dict], brands: list[dict]) -> None:
    cleaned = _cleaned_rows(rows)
    for variant, fill in (("working", False), ("solution", True)):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"
        write_sales_enriched(ws, cleaned)
        add_refs(wb, reps, brands)
        ex = wb.create_sheet("Exercises", 0)
        add_note(ex, "Stage 3 — explore & summarize. 'Sales' is enriched (SalesManager, "
                     "BrandManager via VLOOKUP). Filter with Ctrl+Shift+L and read the Status Bar, "
                     "then put each formula in column C.")
        task_header(ex, 3, ["#", "Question", "Your formula"], [4, 74, 60])
        for i, (label, sol, fmt) in enumerate(STAGE3_PROMPTS, start=1):
            row = 3 + i
            ex.cell(row=row, column=1, value=i)
            ex.cell(row=row, column=2, value=label)
            if fill:
                ex.cell(row=row, column=3, value=sol).number_format = fmt
        out = (SOLUTIONS_DIR if variant == "solution" else WORKING_DIR) / "module-3.xlsx"
        wb.save(out)


# ===================== Stage 4 · Pivot & rank ===============================

def build_stage4(rows: list[dict], reps: list[dict], brands: list[dict]) -> None:
    cleaned = _cleaned_rows(rows)

    # ---- working ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    write_sales_enriched(ws, cleaned)
    add_refs(wb, reps, brands)
    pad = wb.create_sheet("Pivot Scratchpad", 0)
    add_note(pad, "Stage 4 — build PivotTables here (Insert > PivotTable). Add Slicers for Area "
                  "and InvoiceType. Last two tasks are the leaderboards.")
    task_list(pad, [
        "Pivot 1 — Net sales by Area (Rows = Area, Values = Sum of SalesValue).",
        "Pivot 2 — Net sales by Category x Quarter (group Date by Quarter into Columns).",
        "Pivot 3 — Net sales by Customer, with an InvoiceType slicer.",
        "Pivot 4 — % of Grand Total by Brand (Show Values As > % of Grand Total).",
        "Leaderboard A — Top 5 Sales Reps by net sales (Value Filter > Top 10 > 5), sorted.",
        "Leaderboard B — Net sales by Brand Manager (Rows = BrandManager).",
    ])
    wb.save(WORKING_DIR / "module-4.xlsx")

    # ---- solution ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    write_sales_enriched(ws, cleaned)
    add_refs(wb, reps, brands)
    sm = wb.create_sheet("Summary", 0)
    add_note(sm, "Stage 4 solution. PivotTables/Slicers can't be authored by openpyxl, so these "
                 "SUMIFS tables mirror what your pivots should show.")
    areas = sorted({r["Area"].title() for r in cleaned})

    section(sm, 3, "Net sales by Area")
    sm["A4"] = "Area"; sm["B4"] = "Net Sales"
    for c in ("A4", "B4"):
        sm[c].font = HEADER_FONT; sm[c].fill = HEADER_FILL
    for i, a in enumerate(areas, start=5):
        sm.cell(row=i, column=1, value=a)
        sm.cell(row=i, column=2, value=f'=SUMIFS(Sales[SalesValue],Sales[Area],A{i})').number_format = AED

    # Sales rep leaderboard (net sales) with data bars
    start = 5 + len(areas) + 2
    section(sm, start - 1, "Sales Rep Leaderboard  (net sales, ranked, data bars)")
    sm.cell(row=start, column=1, value="SalesRep").font = HEADER_FONT
    sm.cell(row=start, column=1).fill = HEADER_FILL
    sm.cell(row=start, column=2, value="Net Sales").font = HEADER_FONT
    sm.cell(row=start, column=2).fill = HEADER_FILL
    rep_tot: dict[str, float] = {}
    for r in cleaned:
        rep_tot[r["SalesRep"]] = rep_tot.get(r["SalesRep"], 0.0) + r["SalesValue"]
    ranked = sorted(rep_tot.items(), key=lambda kv: -kv[1])
    for i, (rep, _t) in enumerate(ranked, start=start + 1):
        sm.cell(row=i, column=1, value=rep)
        sm.cell(row=i, column=2,
                value=f'=SUMIFS(Sales[SalesValue],Sales[SalesRep],A{i})').number_format = AED
    sm.conditional_formatting.add(f"B{start+1}:B{start+len(ranked)}",
                                  DataBarRule(start_type="min", end_type="max", color="638EC6"))

    # Brand manager leaderboard
    bm_start = start + len(ranked) + 2
    section(sm, bm_start - 1, "Net sales by Brand Manager")
    sm.cell(row=bm_start, column=1, value="BrandManager").font = HEADER_FONT
    sm.cell(row=bm_start, column=1).fill = HEADER_FILL
    sm.cell(row=bm_start, column=2, value="Net Sales").font = HEADER_FONT
    sm.cell(row=bm_start, column=2).fill = HEADER_FILL
    bms = sorted({b["BrandManager"] for b in brands})
    for i, bm in enumerate(bms, start=bm_start + 1):
        sm.cell(row=i, column=1, value=bm)
        sm.cell(row=i, column=2,
                value=f'=SUMIFS(Sales[SalesValue],Sales[BrandManager],A{i})').number_format = AED

    sm.column_dimensions["A"].width = 22
    sm.column_dimensions["B"].width = 16
    wb.save(SOLUTIONS_DIR / "module-4.xlsx")


# ===================== Stage 5 · Present ====================================

def build_stage5(rows: list[dict], reps: list[dict], brands: list[dict]) -> None:
    cleaned = _cleaned_rows(rows)

    # ---- working ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    write_sales_enriched(ws, cleaned)
    add_refs(wb, reps, brands)
    pg = wb.create_sheet("One-Pager", 0)
    add_note(pg, "Stage 5 — build a one-page summary for the VP on this sheet. Keep it to one "
                 "screen, format AED cleanly, and let conditional formatting do the flagging.")
    task_list(pg, [
        "KPI strip: Net sales, Gross sales, Return rate, Average sale value, # invoices.",
        "One chart: a column chart of net sales by Area (sorted, titled, AED axis).",
        "Heat map: net sales by Category x Area with a colour scale (Conditional Formatting).",
        "Data bars on a brand-net-sales column for an instant ranking.",
        "Flag rules: highlight duplicate Customer names; red-fill rows where InvoiceType = Return.",
        "Format AED as #,##0 and fit everything on one screen at 100%.",
    ])
    wb.save(WORKING_DIR / "module-5.xlsx")

    # ---- solution ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    write_sales_enriched(ws, cleaned)
    add_refs(wb, reps, brands)
    dash = wb.create_sheet("One-Pager", 0)
    add_note(dash, "Stage 5 solution — a one-page summary: KPI strip + a net-sales-by-Area chart "
                   "+ a Category x Area heat map, all from live SUMIFS.")
    kpis = [
        ("Net Sales", "=SUM(Sales[SalesValue])", AED),
        ("Gross Sales", '=SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Sales")', AED),
        ("Return Rate",
         '=-SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Return")/SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Sales")', "0.0%"),
        ("Avg Sale Value", '=AVERAGEIFS(Sales[SalesValue],Sales[InvoiceType],"Sales")', AED),
        ("# Invoices", "=COUNTA(Sales[OrderNumber])", "#,##0"),
    ]
    for i, (label, formula, fmt) in enumerate(kpis):
        col = 1 + i * 2
        dash.cell(row=3, column=col, value=label).font = Font(bold=True, color="1F4E78")
        c = dash.cell(row=4, column=col, value=formula)
        c.font = Font(bold=True, size=14)
        c.number_format = fmt
        dash.column_dimensions[get_column_letter(col)].width = 18

    section(dash, 7, "Net sales by Area")
    dash["A8"] = "Area"; dash["B8"] = "Net Sales"
    for c in ("A8", "B8"):
        dash[c].font = HEADER_FONT; dash[c].fill = HEADER_FILL
    areas = sorted({r["Area"].title() for r in cleaned})
    for i, a in enumerate(areas, start=9):
        dash.cell(row=i, column=1, value=a)
        dash.cell(row=i, column=2, value=f'=SUMIFS(Sales[SalesValue],Sales[Area],A{i})').number_format = AED
    chart = BarChart(); chart.type = "col"; chart.title = "Net Sales by Area"
    data = Reference(dash, min_col=2, min_row=8, max_row=8 + len(areas))
    cats = Reference(dash, min_col=1, min_row=9, max_row=8 + len(areas))
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    chart.height = 8; chart.width = 16
    dash.add_chart(chart, "D7")

    base = 9 + len(areas) + 2
    section(dash, base - 1, "Net sales by Category x Area  (colour scale = heat map)")
    dash.cell(row=base, column=1, value="Category").font = HEADER_FONT
    dash.cell(row=base, column=1).fill = HEADER_FILL
    for j, a in enumerate(areas, start=2):
        c = dash.cell(row=base, column=j, value=a)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    for i, cat in enumerate(["Food", "HPC"], start=base + 1):
        dash.cell(row=i, column=1, value=cat)
        for j, a in enumerate(areas, start=2):
            col = get_column_letter(j)
            dash.cell(row=i, column=j,
                      value=f'=SUMIFS(Sales[SalesValue],Sales[Category],$A{i},Sales[Area],{col}${base})').number_format = AED
    end_col = get_column_letter(1 + len(areas))
    dash.conditional_formatting.add(
        f"B{base+1}:{end_col}{base+2}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))
    dash.column_dimensions["A"].width = 14
    wb.save(SOLUTIONS_DIR / "module-5.xlsx")


# =============================== Capstone ===================================

CAPSTONE_QUESTIONS: list[tuple[str, str, str]] = [
    ("1. Net sales (sum of SalesValue)?", "=SUM(Sales[SalesValue])", AED),
    ("2. Gross sales (Sales invoices only)?",
     '=SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Sales")', AED),
    ("3. Total returns value (Return invoices)?",
     '=SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Return")', AED),
    ("4. Return rate (returns / gross, positive %)?",
     '=-SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Return")'
     '/SUMIFS(Sales[SalesValue],Sales[InvoiceType],"Sales")', "0.0%"),
    ("5. Number of invoice lines?", "=COUNTA(Sales[OrderNumber])", "#,##0"),
    ("6. Average sale value (Sales lines only)?",
     '=AVERAGEIFS(Sales[SalesValue],Sales[InvoiceType],"Sales")', AED),
    (f"7. Net sales in the {EG_AREA} area?",
     f'=SUMIFS(Sales[SalesValue],Sales[Area],"{EG_AREA}")', AED),
    (f"8. Net sales for the {EG_CATEGORY} category?",
     f'=SUMIFS(Sales[SalesValue],Sales[Category],"{EG_CATEGORY}")', AED),
    (f"9. Net sales for customer {EG_CUSTOMER}?",
     f'=SUMIFS(Sales[SalesValue],Sales[Customer],"{EG_CUSTOMER}")', AED),
    (f"10. Net sales for brand manager {EG_BRANDMGR}? (BrandManager column)",
     f'=SUMIFS(Sales[SalesValue],Sales[BrandManager],"{EG_BRANDMGR}")', AED),
    (f"11. {EG_REP}'s quota attainment (net sales / quota from Reps)?",
     f'=SUMIFS(Sales[SalesValue],Sales[SalesRep],"{EG_REP}")/VLOOKUP("{EG_REP}",Reps[#All],3,FALSE)', "0.0%"),
]


def build_capstone(rows: list[dict], reps: list[dict], brands: list[dict]) -> None:
    cleaned = _cleaned_rows(rows)

    def base(wb):
        ws = wb.active
        ws.title = "Sales"
        write_sales_enriched(ws, cleaned)
        add_refs(wb, reps, brands)

    wb = Workbook()
    base(wb)
    q = wb.create_sheet("Questions", 0)
    add_note(q, "CAPSTONE — a fresh file. The 'Sales' table is enriched (SalesManager, "
                "BrandManager). Answer each question in column C using PivotTables, AutoFilter, "
                "SUMIFS/COUNTIFS and VLOOKUP — whatever is fastest.")
    task_header(q, 3, ["#", "Business question", "Your answer"], [4, 80, 24])
    for i, (question, _sol, _fmt) in enumerate(CAPSTONE_QUESTIONS, start=1):
        q.cell(row=3 + i, column=1, value=i)
        q.cell(row=3 + i, column=2, value=question)
        q.cell(row=3 + i, column=3).fill = TASK_FILL
    wb.save(WORKING_DIR / "capstone.xlsx")

    wb = Workbook()
    base(wb)
    a = wb.create_sheet("Answers", 0)
    add_note(a, "CAPSTONE answer key. Each answer is a live formula so it stays correct if the "
                "data changes. Many also have a valid PivotTable route.")
    task_header(a, 3, ["#", "Business question", "Answer"], [4, 80, 24])
    for i, (question, sol, fmt) in enumerate(CAPSTONE_QUESTIONS, start=1):
        a.cell(row=3 + i, column=1, value=i)
        a.cell(row=3 + i, column=2, value=question)
        a.cell(row=3 + i, column=3, value=sol).number_format = fmt
    wb.save(SOLUTIONS_DIR / "capstone.xlsx")


# ------------------------------ entry ---------------------------------------

def main() -> None:
    WORKING_DIR.mkdir(parents=True, exist_ok=True)
    SOLUTIONS_DIR.mkdir(parents=True, exist_ok=True)

    rows = load_rows()
    reps = load_reps()
    brands = load_brands()
    print(f"Loaded {len(rows)} rows, {len(reps)} reps, {len(brands)} brands")

    build_stage1(rows);                  print("  stage 1 (module-1) done")
    build_stage2(rows, reps, brands);    print("  stage 2 (module-2) done")
    build_stage3(rows, reps, brands);    print("  stage 3 (module-3) done")
    build_stage4(rows, reps, brands);    print("  stage 4 (module-4) done")
    build_stage5(rows, reps, brands);    print("  stage 5 (module-5) done")
    build_capstone(rows, reps, brands);  print("  capstone done")


if __name__ == "__main__":
    main()
